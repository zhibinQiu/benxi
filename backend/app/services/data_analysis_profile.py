"""表格分析 — 表格结构画像（Excel / CSV，不向 LLM 传递全量数据）。"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from app.core.exceptions import bad_request

logger = logging.getLogger(__name__)

SAMPLE_ROWS = 5
PROFILE_SAMPLE_ROWS = 200
CSV_SHEET_NAME = "data"

_DATA_ANALYSIS_MODULES = (
    "pandas",
    "openpyxl",
    "numpy",
    "matplotlib",
    "seaborn",
)


def missing_data_analysis_modules() -> list[str]:
    missing: list[str] = []
    for name in _DATA_ANALYSIS_MODULES:
        try:
            __import__(name)
        except ImportError:
            missing.append(name)
    return missing


def data_analysis_deps_install_hint() -> str:
    return (
        "本地开发: cd platform && pip install -e . ；"
        "Docker 栈: bash scripts/stack.sh build && bash scripts/stack.sh dev-up（或 up）"
    )


def warn_if_data_analysis_deps_missing() -> None:
    missing = missing_data_analysis_modules()
    if missing:
        logger.warning(
            "表格分析依赖未就绪 (%s)。%s",
            ", ".join(missing),
            data_analysis_deps_install_hint(),
        )


def _require_pandas():
    try:
        import pandas as pd
        from openpyxl import load_workbook
    except ImportError as exc:
        missing = missing_data_analysis_modules()
        detail = ", ".join(missing) if missing else "pandas/openpyxl"
        raise bad_request(
            f"表格分析依赖未安装（{detail}）。{data_analysis_deps_install_hint()}"
        ) from exc
    return pd, load_workbook


def detect_file_type(path: Path, *, filename: str = "") -> str:
    ext = Path(filename or path.name).suffix.lower()
    if ext == ".csv":
        return "csv"
    if ext in {".xlsx", ".xls"}:
        return "excel"
    if path.suffix.lower() == ".csv":
        return "csv"
    return "excel"


def _fmt_value(value: Any) -> str:
    pd, _ = _require_pandas()
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    text = str(value)
    return text if len(text) <= 80 else text[:77] + "..."


def _column_profile(series) -> dict[str, Any]:
    pd, _ = _require_pandas()
    non_null = series.dropna()
    profile: dict[str, Any] = {
        "name": str(series.name),
        "dtype": str(series.dtype),
        "null_count": int(series.isna().sum()),
        "non_null_count": int(non_null.shape[0]),
        "sample_values": [_fmt_value(v) for v in non_null.head(5).tolist()],
    }
    if pd.api.types.is_numeric_dtype(series):
        if non_null.empty:
            return profile
        profile["min_value"] = _fmt_value(non_null.min())
        profile["max_value"] = _fmt_value(non_null.max())
        profile["mean_value"] = _fmt_value(non_null.mean())
    return profile


def _sheet_profile_from_frame(sample, *, name: str, row_count: int) -> dict[str, Any]:
    sample_rows = [
        [_fmt_value(v) for v in row.tolist()]
        for _, row in sample.head(SAMPLE_ROWS).iterrows()
    ]
    return {
        "name": name,
        "rows": row_count,
        "columns": int(sample.shape[1]),
        "column_profiles": [_column_profile(sample[col]) for col in sample.columns],
        "sample_rows": sample_rows,
    }


def _csv_row_count(path: Path) -> int:
    with path.open("r", encoding="utf-8", errors="replace") as handle:
        total = sum(1 for _ in handle)
    return max(total - 1, 0)


def _sheet_row_count(path: Path, sheet_name: str) -> int:
    _, load_workbook = _require_pandas()
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return 0
        ws = wb[sheet_name]
        return max(int(ws.max_row or 0) - 1, 0)
    finally:
        wb.close()


def build_csv_profile(path: Path, *, filename: str) -> dict[str, Any]:
    pd, _ = _require_pandas()
    sample = pd.read_csv(path, nrows=PROFILE_SAMPLE_ROWS)
    row_count = _csv_row_count(path)
    sheet = _sheet_profile_from_frame(sample, name=CSV_SHEET_NAME, row_count=row_count)
    return {
        "filename": filename,
        "file_type": "csv",
        "sheets": [sheet],
        "active_sheet": CSV_SHEET_NAME,
        "file_size_bytes": path.stat().st_size,
    }


def build_excel_profile(path: Path, *, filename: str) -> dict[str, Any]:
    pd, _ = _require_pandas()
    xl = pd.ExcelFile(path)
    sheets: list[dict[str, Any]] = []
    for sheet_name in xl.sheet_names:
        row_count = _sheet_row_count(path, sheet_name)
        sample = pd.read_excel(path, sheet_name=sheet_name, nrows=PROFILE_SAMPLE_ROWS)
        sheets.append(_sheet_profile_from_frame(sample, name=sheet_name, row_count=row_count))
    active = xl.sheet_names[0] if xl.sheet_names else ""
    return {
        "filename": filename,
        "file_type": "excel",
        "sheets": sheets,
        "active_sheet": active,
        "file_size_bytes": path.stat().st_size,
    }


def build_dataset_profile(path: Path, *, filename: str) -> dict[str, Any]:
    if detect_file_type(path, filename=filename) == "csv":
        return build_csv_profile(path, filename=filename)
    return build_excel_profile(path, filename=filename)


def profile_summary_for_llm(profile: dict[str, Any]) -> str:
    file_type = profile.get("file_type") or "excel"
    lines = [
        f"文件名: {profile.get('filename', '')}",
        f"文件类型: {'CSV' if file_type == 'csv' else 'Excel'}",
    ]
    if file_type == "excel":
        lines.append(f"默认工作表: {profile.get('active_sheet', '')}")
    for sheet in profile.get("sheets") or []:
        lines.append("")
        label = "数据表" if file_type == "csv" else f"工作表「{sheet.get('name')}」"
        lines.append(f"## {label}")
        lines.append(f"- 行数(约): {sheet.get('rows')}")
        lines.append(f"- 列数: {sheet.get('columns')}")
        lines.append("- 字段结构:")
        for col in sheet.get("column_profiles") or []:
            parts = [f"  · {col.get('name')} ({col.get('dtype')})"]
            if col.get("null_count") is not None:
                parts.append(f"空值={col.get('null_count')}")
            if col.get("min_value") is not None:
                parts.append(f"min={col.get('min_value')}")
            if col.get("max_value") is not None:
                parts.append(f"max={col.get('max_value')}")
            if col.get("mean_value") is not None:
                parts.append(f"mean={col.get('mean_value')}")
            samples = col.get("sample_values") or []
            if samples:
                parts.append(f"样例={', '.join(samples[:3])}")
            lines.append(", ".join(parts))
        sample_rows = sheet.get("sample_rows") or []
        if sample_rows:
            lines.append("- 前几行样例(仅结构预览):")
            for row in sample_rows[:3]:
                lines.append(f"  | {' | '.join(row)} |")
    lines.append("")
    if file_type == "csv":
        lines.append(
            "运行时已有变量: df (CSV 已加载为 DataFrame), DATA_PATH, FILE_TYPE='csv', "
            "pd, np, plt, sns。"
        )
    else:
        lines.append(
            "运行时已有变量: df (默认工作表 DataFrame), DATA_PATH, FILE_TYPE='excel', "
            "ACTIVE_SHEET, pd, np, plt, sns。切换工作表请 pd.read_excel(DATA_PATH, sheet_name=...)。"
        )
    return "\n".join(lines)
